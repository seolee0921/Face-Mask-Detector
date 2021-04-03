import cv2
from openpyxl import Workbook

image = cv2.imread("teemo minimalistic.jpg")
wb = Workbook()
ws = wb.active
ws = wb.create_sheet("sample")

n = 1
for y in range(1, image.shape[0]):
    for x in range(1, image.shape[1]):
        ws.cell(row=y, column=x, value=n)
        n += 1

wb.save("sample.xlsx")
